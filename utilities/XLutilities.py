import openpyxl


def getRowCount(file, sheetname):
    Book = openpyxl.load_workbook(file)
    Sheet = Book[sheetname]
    return Sheet.max_row


def readData(file, sheetname, rownum, colnum):
    Book = openpyxl.load_workbook(file)
    Sheet = Book[sheetname]
    return Sheet.cell(row=rownum, column=colnum).value


def writeData(file, sheetname, rownum, colnum, data):
    Book = openpyxl.load_workbook(file)
    Sheet = Book[sheetname]
    Sheet.cell(row=rownum, column=colnum).value = data
    Book.save(file)